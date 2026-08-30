from __future__ import annotations

import csv
import hashlib
import json
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config import Settings
from .db import ContextStore, stable_id, stable_json
from .ingest import file_sha256
from .normalization import normalize_vehicle_registration, normalized_text
from .redaction import PIIRedactor


def _csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _source(store: ContextStore, settings: Settings, path: Path, count: int | None) -> str:
    return store.upsert_source(str(path.relative_to(settings.corpus_dir)).replace("\\", "/"), file_sha256(path), count)


def _plus_six_months(value: str) -> str:
    joined = date.fromisoformat(value)
    month = joined.month + 6
    year = joined.year + (month - 1) // 12
    month = (month - 1) % 12 + 1
    # The roster dates never require day clamping, but this keeps the operation valid.
    day = min(joined.day, (date(year + (month == 12), month % 12 + 1, 1) - date.resolution).day)
    return date(year, month, day).isoformat()


def ingest_drivers(settings: Settings, store: ContextStore) -> list[str]:
    path = settings.corpus_dir / "drivers_roster.csv"
    rows = _csv_rows(path)
    source_id = _source(store, settings, path, len(rows))
    names: list[str] = []
    with store.transaction() as conn:
        for index, row in enumerate(rows, start=2):
            if row.get("name"):
                names.append(row["name"])
            driver_id = normalized_text(row.get("driver_id"))
            if not driver_id or not row.get("joining_date"):
                continue
            citation = f"{source_id}:row:{index}"
            conn.execute(
                """INSERT INTO drivers(driver_id, home_hub, night_solo_eligible_after, source_citation)
                   VALUES (?, ?, ?, ?)
                   ON CONFLICT(driver_id) DO UPDATE SET home_hub=excluded.home_hub,
                     night_solo_eligible_after=excluded.night_solo_eligible_after,
                     source_citation=excluded.source_citation""",
                (driver_id, normalized_text(row.get("home_hub")), _plus_six_months(row["joining_date"]), citation),
            )
    store.upsert_document(source_id, "driver_roster", "structured_context", "Redacted driver eligibility roster")
    return names


def ingest_fleet(settings: Settings, store: ContextStore) -> dict[str, int]:
    path = settings.corpus_dir / "fleet_master.csv"
    rows = _csv_rows(path)
    source_id = _source(store, settings, path, len(rows))
    claims: dict[str, list[dict[str, Any]]] = defaultdict(list)
    with store.transaction() as conn:
        for index, row in enumerate(rows, start=2):
            vehicle_reg = normalize_vehicle_registration(row.get("registration_number"))
            if not vehicle_reg:
                continue
            claim = {
                "vehicle_id": normalized_text(row.get("vehicle_id")),
                "model": normalized_text(row.get("model")),
                "year": int(row["year"]) if row.get("year", "").isdigit() else None,
                "bs_stage": normalized_text(row.get("bs_stage")),
                "engine_heater": normalized_text(row.get("engine_heater")),
                "home_hub": normalized_text(row.get("home_hub")),
                "capacity_tonnes": float(row["capacity_tonnes"]) if row.get("capacity_tonnes") else None,
                "fleet_status": normalized_text(row.get("status")),
                "citation": f"{source_id}:row:{index}",
            }
            claims[vehicle_reg].append(claim)
            conn.execute(
                """INSERT INTO vehicle_claims(claim_id, vehicle_reg, source_id, record_index, claim_json, source_priority)
                   VALUES (?, ?, ?, ?, ?, 1)
                   ON CONFLICT(source_id, record_index) DO NOTHING""",
                (stable_id("VCL", source_id, index), vehicle_reg, source_id, index, stable_json(claim)),
            )

    conflict_count = 0
    with store.transaction() as conn:
        for vehicle_reg, vehicle_claims in sorted(claims.items()):
            # Fleet row with a stable vehicle_id is authoritative over the deliberately incomplete sync copy.
            preferred = next((claim for claim in vehicle_claims if claim["vehicle_id"]), vehicle_claims[0])
            material_conflict = False
            for field in ("year", "bs_stage", "engine_heater", "home_hub", "capacity_tonnes", "fleet_status"):
                nonempty = sorted({claim[field] for claim in vehicle_claims if claim[field] not in (None, "")})
                if len(nonempty) > 1:
                    material = field in {"year", "bs_stage", "engine_heater", "home_hub", "fleet_status"}
                    material_conflict = material_conflict or material
                    conflict_count += 1
            conn.execute(
                """INSERT INTO vehicles(vehicle_reg, vehicle_id, model, year, bs_stage, engine_heater,
                   home_hub, capacity_tonnes, fleet_status, resolution_status, source_citation)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT(vehicle_reg) DO UPDATE SET vehicle_id=excluded.vehicle_id, model=excluded.model,
                     year=excluded.year, bs_stage=excluded.bs_stage, engine_heater=excluded.engine_heater,
                     home_hub=excluded.home_hub, capacity_tonnes=excluded.capacity_tonnes,
                     fleet_status=excluded.fleet_status, resolution_status=excluded.resolution_status,
                     source_citation=excluded.source_citation""",
                (vehicle_reg, preferred["vehicle_id"], preferred["model"], preferred["year"], preferred["bs_stage"],
                 preferred["engine_heater"], preferred["home_hub"], preferred["capacity_tonnes"],
                 preferred["fleet_status"], "CONFLICT" if material_conflict else "RESOLVED", preferred["citation"]),
            )
    # Conflict records use their own short transactions; do not nest a second connection
    # inside the fleet-write transaction above.
    for vehicle_reg, vehicle_claims in sorted(claims.items()):
        for field in ("year", "bs_stage", "engine_heater", "home_hub", "capacity_tonnes", "fleet_status"):
            nonempty = sorted({claim[field] for claim in vehicle_claims if claim[field] not in (None, "")})
            if len(nonempty) > 1:
                material = field in {"year", "bs_stage", "engine_heater", "home_hub", "fleet_status"}
                store.upsert_conflict(
                    "vehicle", vehicle_reg, field,
                    [{"value": claim[field], "citation": claim["citation"]} for claim in vehicle_claims],
                    material, [claim["citation"] for claim in vehicle_claims],
                )
    store.upsert_document(source_id, "fleet_master", "structured_context", "Canonical vehicle claims and conflicts")
    return {"fleet_rows": len(rows), "vehicles": len(claims), "conflicts": conflict_count}


def ingest_maintenance(settings: Settings, store: ContextStore, redactor: PIIRedactor) -> int:
    path = settings.corpus_dir / "maintenance_log.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Maintenance Log"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    source_id = _source(store, settings, path, len(rows))
    with store.transaction() as conn:
        for row_index, row in enumerate(rows, start=2):
            event_date, vehicle, odometer, _mechanic, notes = row
            vehicle_reg = normalize_vehicle_registration(vehicle)
            if not vehicle_reg or event_date in (None, "") or odometer is None:
                continue
            citation = f"{source_id}:sheet:Maintenance Log:row:{row_index}"
            text = str(notes or "").lower()
            # Do not persist mixed-language mechanic prose. The deterministic flags below are
            # sufficient for the declared safety rules and cannot carry a person's identity.
            maintenance_features = ",".join(sorted(set(
                feature for feature, matched in {
                    "brake_work": "brake" in text,
                    "temporary_fix": "temporary fix" in text or "jugaad" in text,
                    "permanent_repair_pending": "permanent fix" in text or "needs permanent repair" in text,
                }.items() if matched
            ))) or "no_rule_relevant_feature"
            conn.execute(
                """INSERT INTO maintenance_events(maintenance_id, vehicle_reg, event_date, odometer_km,
                   notes_redacted, source_citation) VALUES (?, ?, ?, ?, ?, ?)
                   ON CONFLICT(source_citation) DO UPDATE SET notes_redacted=excluded.notes_redacted,
                     event_date=excluded.event_date, odometer_km=excluded.odometer_km""",
                (stable_id("MNT", citation), vehicle_reg, str(event_date), float(odometer),
                 maintenance_features, citation),
            )
    workbook.close()
    store.upsert_document(source_id, "maintenance_log", "structured_context", "Redacted maintenance events")
    return len(rows)


def ingest_trips(settings: Settings, store: ContextStore) -> int:
    path = settings.corpus_dir / "meridian_trips.csv"
    rows = _csv_rows(path)
    source_id = _source(store, settings, path, len(rows))
    with store.transaction() as conn:
        for index, row in enumerate(rows, start=2):
            trip_id = normalized_text(row.get("trip_id"))
            if not trip_id:
                continue
            citation = f"{source_id}:row:{index}"
            conn.execute(
                """INSERT INTO trips(trip_id, created_at, route_type, origin_center, origin_name, dest_center,
                   dest_name, dispatch_time, delivery_time, osrm_distance_km, osrm_time_min, actual_time_min,
                   vehicle_reg, driver_id, client, trip_status, billed_amount, source_citation)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT(trip_id) DO NOTHING""",
                (trip_id, row["created_at"], row["route_type"], row["origin_center"], normalized_text(row.get("origin_name")),
                 row["dest_center"], normalized_text(row.get("dest_name")), row["dispatch_time"], row["delivery_time"],
                 float(row["osrm_distance_km"]), float(row["osrm_time_min"]), float(row["actual_time_min"]),
                 normalize_vehicle_registration(row["vehicle_reg"]) or "", row["driver_id"], row["client"], row["status"],
                 float(row["billed_amount"]), citation),
            )
    store.upsert_document(source_id, "trip_history", "historical_only", "Historical trips; not current location or availability")
    return len(rows)


def ingest_documents(settings: Settings, store: ContextStore) -> int:
    paths = [settings.corpus_dir / "dispatcher_interview.txt"] + sorted((settings.corpus_dir / "emails").glob("*.txt"))
    count = 0
    for path in paths:
        source_id = _source(store, settings, path, 1)
        if path.name.startswith("thread_"):
            number = int(path.name.split("_")[1])
            classification = "operational_evidence" if number <= 25 else "non_operational_noise"
            document_type = "email_thread"
        else:
            classification = "authoritative_rules"
            document_type = "dispatcher_interview"
        # Do not persist email/transcript prose: it contains unbounded personal data. Rules are
        # later stored as structured, redacted facts with source-only citations.
        store.upsert_document(source_id, document_type, classification, "Content held only in source file; cite by source ID")
        count += 1
    return count


def ingest_context(settings: Settings, store: ContextStore) -> dict[str, int]:
    names = ingest_drivers(settings, store)
    redactor = PIIRedactor(names)
    fleet = ingest_fleet(settings, store)
    maintenance_count = ingest_maintenance(settings, store, redactor)
    trip_count = ingest_trips(settings, store)
    document_count = ingest_documents(settings, store)
    return {**fleet, "drivers": len(names), "maintenance": maintenance_count, "trips": trip_count, "documents": document_count}
